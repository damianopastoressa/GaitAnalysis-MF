import openpyxl

class FeaturesReader:
    
    def __init__(self, file_path):
        self.file_path = file_path
        self.features = []
        self.labels = []
        self.coordL = self.enumerator()
        
        

    #function to enumerate excel column needed to read features
    def enumerator(self):
        coordL = []
        alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
        alphabetPlus = [''] + alphabet
        count = 0
        for j in range(0, len(alphabetPlus)):
            for k in range(0, len(alphabetPlus)):
                for i in range(0, len(alphabet)):
                    coordL.append(alphabetPlus[j]+alphabetPlus[k]+alphabet[i])
                    count += 1
        return coordL        
        


    # function to read features for train phase
    def train_reader(self):
        file = openpyxl.load_workbook(self.file_path)
        sheet = file.worksheets[0]
        for row in range(0, sheet.max_row):
            tup = []
            for i in range (0, sheet.max_column):
                pos = str(self.coordL[i]) + str(row + 1)
                if (i != sheet.max_column-1):
                    tup.append(sheet[pos].value)
                else:
                    label = sheet[pos].value
            if(len(tup) > 1):
                self.features.append(tup)
                self.labels.append(label)
        file.close
        return self.features, self.labels
    
    
    
    # function to read features for test phase        
    def test_reader(self):
        file = openpyxl.load_workbook(self.file_path)
        sheet = file.worksheets[0]
        for row in range(0, sheet.max_row):
            tup = []
            for i in range (0, sheet.max_column):
                pos = str(self.coordL[i]) + str(row + 1)
                tup.append(sheet[pos].value)
            if(len(tup) > 1):
                self.features.append(tup)
        file.close
        return self.features